#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Dashboard วิเคราะห์ผลการสำรวจสภาวะช่องปาก (HDC Data Exchange)
============================================================
รวมไฟล์ Excel จาก data/<กลุ่มอายุ>/*.xlsx  →  คำนวณตารางที่ 1-7  →  สร้าง dashboard.html

ใช้ได้ทั้งแบบรัน Python ตรง ๆ และแบบแพ็กเป็นไฟล์ .exe ด้วย PyInstaller

    python src/build_dashboard.py            (โหมดพัฒนา)
    Dashboardช่องปาก.exe                      (ผู้ใช้ปลายทาง ดับเบิลคลิก)

ตัวเลือกเพิ่มเติม
    --force        ประมวลผลไฟล์ Excel ใหม่ทั้งหมด ไม่ใช้ข้อมูลที่แคชไว้
    --full-csv     เก็บทุกคอลัมน์ลง combined.csv รวมชื่อ สกุล เลขบัตร ที่อยู่
    --no-open      ไม่ต้องเปิดเบราว์เซอร์เมื่อเสร็จ
"""

import csv
import json
import os
import re
import shutil
import sys
import tempfile
import traceback
import webbrowser
from collections import defaultdict
from datetime import datetime
from pathlib import Path

APP_NAME = "Dashboard วิเคราะห์สภาวะช่องปาก"
VERSION = "1.1.2"

# ที่อยู่ repository บน GitHub ใช้สำหรับแจ้งเตือนเมื่อมีเวอร์ชันใหม่ เช่น "myname/ohsp-data-ex"
# ตอน build ผ่าน GitHub Actions หรือ build.bat ค่านี้จะถูกเติมให้อัตโนมัติจาก remote ของ repo
# ถ้าเว้นว่างไว้ โปรแกรมจะข้ามการตรวจสอบเวอร์ชันไปเลย
GITHUB_REPO = "nchotenin-tech/ohsp-data-ex"

AGE_ORDER = ["3ปี", "6ปี", "12ปี", "60ปี"]
AGE_LABEL = {"3ปี": "3", "6ปี": "6", "12ปี": "12", "60ปี": "60+"}
AGE_BY_NUMBER = {3: "3ปี", 6: "6ปี", 12: "12ปี", 60: "60ปี"}
NA_VALUES = {"", "<na>", "na", "nan", "none", "null", "-"}

# คอลัมน์ที่ต้องมีในไฟล์ Excel มิฉะนั้นคำนวณไม่ได้
REQUIRED_COLS = [
    "hoscode", "date_serv", "providertype",
    "pteeth", "pcaries", "pfilling", "pextract",
    "dteeth", "dcaries", "dfilling", "dextract",
    "need_fluoride", "need_scaling", "need_sealant",
    "need_pfilling", "need_dfilling", "need_pextract", "need_dextract",
    "permanent_permanent", "permanent_prosthesis", "prosthesis_prosthesis", "gum",
]
# คอลัมน์เสริมที่เก็บไว้ด้วยถ้ามี (ไม่ใช่ข้อมูลระบุตัวบุคคล)
EXTRA_COLS = ["hosname", "pid", "seq", "age_y", "sex", "denttype", "servplace",
              "nprosthesis", "class", "schooltype", "provider"]
# คอลัมน์ข้อมูลส่วนบุคคล ตัดออกจาก combined.csv เว้นแต่สั่ง --full-csv
PII_COLS = ["name", "lname", "cid", "addr", "birth", "check_vhid"]


# ---------------------------------------------------------------------------
# เตรียมสภาพแวดล้อม (รองรับทั้งโหมด .py และ .exe)
# ---------------------------------------------------------------------------
FROZEN = getattr(sys, "frozen", False)
if FROZEN:
    APP_DIR = Path(sys.executable).resolve().parent      # โฟลเดอร์ที่วางไฟล์ .exe
    RES_DIR = Path(getattr(sys, "_MEIPASS", APP_DIR))    # ไฟล์ที่ฝังมากับ .exe
else:
    APP_DIR = Path(__file__).resolve().parent.parent     # รากของโปรเจกต์
    RES_DIR = Path(__file__).resolve().parent            # โฟลเดอร์ src

DATA_DIR = APP_DIR / "data"
OUT_DIR = APP_DIR / "ผลลัพธ์"
COMBINED = OUT_DIR / "combined.csv"
META = OUT_DIR / "meta.json"
SUMMARY = OUT_DIR / "summary.json"
OUTPUT = OUT_DIR / "dashboard.html"
LOGFILE = OUT_DIR / "log.txt"


def resource(name):
    """หาไฟล์ประกอบ: ถ้าผู้ใช้วางไว้ข้าง .exe ให้ใช้ของผู้ใช้ก่อน มิฉะนั้นใช้ที่ฝังมา"""
    for base in (APP_DIR, RES_DIR, APP_DIR / "src"):
        p = base / name
        if p.exists():
            return p
    return None


def setup_console():
    """ให้ภาษาไทยแสดงถูกต้องบนหน้าจอ Command Prompt ของ Windows"""
    if os.name == "nt":
        os.system("chcp 65001 >nul 2>&1")
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass


_log_fh = None


def say(msg=""):
    print(msg, flush=True)
    if _log_fh:
        try:
            _log_fh.write(msg + "\n")
            _log_fh.flush()
        except Exception:
            pass


def hr():
    say("─" * 66)


def pause():
    """ค้างหน้าจอไว้ ไม่ให้หน้าต่างปิดวืบเมื่อดับเบิลคลิก .exe
    ข้ามเมื่อไม่ได้รันจากหน้าจอจริง เช่น ในระบบทดสอบอัตโนมัติหรืองานตั้งเวลา
    มิฉะนั้นโปรแกรมจะค้างรอการกดปุ่มที่ไม่มีวันมาถึง"""
    if not FROZEN:
        return
    try:
        if not sys.stdin or not sys.stdin.isatty():
            return
        input("\nกด Enter เพื่อปิดหน้าต่าง ...")
    except Exception:
        pass


# ---------------------------------------------------------------------------
# 1) ค้นหาและรวมไฟล์ Excel
# ---------------------------------------------------------------------------
def age_from_folder(name):
    """แปลงชื่อโฟลเดอร์เป็นกลุ่มอายุ รองรับ '3ปี', '3 ปี', 'age3', '3y', '60+'"""
    nums = re.findall(r"\d+", name)
    for n in nums:
        if int(n) in AGE_BY_NUMBER:
            return AGE_BY_NUMBER[int(n)]
    return None


def ensure_data_folders():
    """สร้างโครงสร้างโฟลเดอร์ให้ผู้ใช้เมื่อเปิดโปรแกรมครั้งแรก"""
    created = not DATA_DIR.exists()
    for age in AGE_ORDER:
        folder = DATA_DIR / age
        folder.mkdir(parents=True, exist_ok=True)
        note = folder / "วางไฟล์ Excel ของกลุ่มอายุนี้ที่นี่.txt"
        if not any(p.suffix.lower() == ".xlsx" for p in folder.iterdir()) and not note.exists():
            note.write_text(
                f"นำไฟล์ Excel ที่ส่งออกจากระบบ HDC (Data Exchange การตรวจฟัน) "
                f"ของกลุ่มอายุ {age} มาวางในโฟลเดอร์นี้\n"
                f"ถ้าข้อมูลถูกแบ่งเป็นหลายไฟล์ ให้วางทุกไฟล์ โปรแกรมจะนำมาต่อกันให้เอง\n",
                encoding="utf-8-sig")
    return created


def find_excel_files():
    """คืนรายการ (กลุ่มอายุ, path) เรียงตามลำดับกลุ่มอายุ"""
    found, unknown = [], []
    for folder in sorted(DATA_DIR.iterdir()):
        if not folder.is_dir():
            continue
        age = age_from_folder(folder.name)
        if age is None:
            unknown.append(folder.name)
            continue
        for f in sorted(folder.glob("*.xlsx")):
            if f.name.startswith("~$"):
                say(f"   ! ข้ามไฟล์ชั่วคราว {f.name} (ไฟล์นี้เกิดจากการเปิดค้างไว้ใน Excel)")
                continue
            found.append((age, f))
    for name in unknown:
        say(f"   ! ข้ามโฟลเดอร์ '{name}' เพราะไม่ทราบว่าเป็นกลุ่มอายุใด "
            f"(ชื่อโฟลเดอร์ต้องมีเลข 3, 6, 12 หรือ 60)")
    found.sort(key=lambda x: (AGE_ORDER.index(x[0]), x[1].name))
    return found


def combine_excels(force=False, full_csv=False):
    import openpyxl                                       # นำเข้าที่นี่เพื่อให้เปิดโปรแกรมได้เร็ว

    files = find_excel_files()
    if not files:
        raise UserError(
            "ไม่พบไฟล์ Excel (.xlsx) ในโฟลเดอร์ data",
            [f"กรุณานำไฟล์ที่ส่งออกจากระบบ HDC มาวางในโฟลเดอร์ต่อไปนี้",
             *[f"    {DATA_DIR / a}" for a in AGE_ORDER],
             "แล้วเปิดโปรแกรมนี้อีกครั้ง"])

    if not force and COMBINED.exists():
        newest = max(p.stat().st_mtime for _, p in files)
        cached_version = None
        try:
            cached_version = json.loads(META.read_text(encoding="utf-8")).get("version")
        except Exception:
            pass
        if cached_version != VERSION:
            # อัปเกรดโปรแกรมมาใหม่ วิธีคำนวณอาจเปลี่ยน จึงต้องอ่านไฟล์ Excel ใหม่ทั้งหมด
            say(f"   อัปเดตโปรแกรมเป็นเวอร์ชัน {VERSION} จึงประมวลผลข้อมูลใหม่ทั้งหมดหนึ่งครั้ง")
        elif COMBINED.stat().st_mtime >= newest:
            say(f"   ใช้ข้อมูลที่ประมวลผลไว้แล้ว ({COMBINED.name}) เพราะไฟล์ Excel ไม่มีการเปลี่ยนแปลง")
            say(f"   ถ้าต้องการประมวลผลใหม่ทั้งหมด ให้ลบไฟล์ {COMBINED.name} แล้วเปิดโปรแกรมอีกครั้ง")
            return

    columns = None
    total = 0
    latest = None
    skipped = []
    # เขียนลงโฟลเดอร์ชั่วคราวก่อน เร็วกว่าและกันไฟล์ค้างครึ่ง ๆ กลาง ๆ เมื่อถูกขัดจังหวะ
    tmp = Path(tempfile.gettempdir()) / f"combined_{os.getpid()}.csv"
    with open(tmp, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        for age, path in files:
            try:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            except Exception as e:
                skipped.append((path.name, f"เปิดไฟล์ไม่ได้ ({type(e).__name__})"))
                say(f"   ! ข้าม {path.name} — เปิดไฟล์ไม่ได้")
                continue

            ws = wb["Data"] if "Data" in wb.sheetnames else wb[wb.sheetnames[0]]
            rows = ws.iter_rows(values_only=True)
            try:
                header = [str(c).strip().lower() if c is not None else "" for c in next(rows)]
            except StopIteration:
                skipped.append((path.name, "ไฟล์ว่าง"))
                wb.close()
                continue

            missing = [c for c in REQUIRED_COLS if c not in header]
            if missing:
                wb.close()
                raise UserError(
                    f"ไฟล์ {path.name} ไม่มีคอลัมน์ที่จำเป็น",
                    [f"คอลัมน์ที่ขาด: {', '.join(missing)}",
                     "กรุณาตรวจสอบว่าเป็นไฟล์ที่ส่งออกจากระบบ HDC (Data Exchange การตรวจฟัน) "
                     "และไม่ได้ถูกแก้ไขหัวตาราง"])

            if columns is None:
                if full_csv:
                    columns = [c for c in header if c]
                else:
                    columns = [c for c in REQUIRED_COLS + EXTRA_COLS if c in header]
                writer.writerow(columns + ["agesurvey"])

            picks = [header.index(c) if c in header else -1 for c in columns]
            date_at = header.index("date_serv")
            width = len(header)
            n = 0
            for r in rows:
                if r is None or all(v is None for v in r):
                    continue
                if len(r) < width:
                    r = tuple(r) + (None,) * (width - len(r))
                writer.writerow(["" if r[i] is None else r[i] for i in picks] + [age])
                d = r[date_at]
                if d is not None:
                    d = str(d).strip()
                    if d.lower() not in NA_VALUES and (latest is None or d > latest):
                        latest = d
                n += 1
                if n % 25000 == 0:
                    say(f"      ... {n:,} แถว")
            wb.close()
            total += n
            say(f"   + {age}/{path.name}  →  {n:,} แถว")

    if columns is None:
        raise UserError("อ่านข้อมูลจากไฟล์ Excel ไม่ได้เลย",
                        [f"{name}: {why}" for name, why in skipped] or ["ไม่ทราบสาเหตุ"])

    shutil.copyfile(str(tmp), str(COMBINED))               # เขียนทับได้โดยไม่ต้องลบไฟล์เดิม
    try:
        tmp.unlink()
    except OSError:
        pass
    META.write_text(json.dumps({"version": VERSION, "latestServe": latest, "rows": total,
                                "fullCsv": full_csv,
                                "files": [p.name for _, p in files]},
                               ensure_ascii=False), encoding="utf-8")
    say(f"   = รวมได้ {total:,} แถว จาก {len(files) - len(skipped)} ไฟล์")
    if not full_csv:
        say(f"   (ไฟล์ {COMBINED.name} ไม่มีชื่อ สกุล เลขบัตรประชาชน และที่อยู่ "
            f"ถ้าต้องการเก็บครบให้เปิดโปรแกรมด้วยตัวเลือก --full-csv)")
    for name, why in skipped:
        say(f"   ! ข้ามไฟล์ {name}: {why}")


# ---------------------------------------------------------------------------
# 2) ตารางเทียบรหัสหน่วยบริการกับอำเภอ/จังหวัด
# ---------------------------------------------------------------------------
def strip_code(text):
    """'01-เมืองชัยภูมิ' -> 'เมืองชัยภูมิ'"""
    return re.sub(r"^\d+\s*-\s*", "", (text or "").strip())


def load_hospitals():
    path = resource("hospitals.csv")
    mapping = {}
    if path is None:
        say("   ! ไม่พบไฟล์ hospitals.csv จะไม่สามารถแยกรายอำเภอได้")
        return mapping
    for enc in ("cp874", "utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(path, encoding=enc, newline="") as fh:
                mapping.clear()
                for row in csv.DictReader(fh):
                    code = (row.get("hoscode") or "").strip()
                    if code:
                        mapping[code] = {
                            "hosname": (row.get("hosname") or "").strip(),
                            "ampname": strip_code(row.get("ampname")),
                            "pvname": strip_code(row.get("pvname")),
                        }
            break
        except UnicodeDecodeError:
            continue
    say(f"   + hospitals.csv: {len(mapping):,} หน่วยบริการ")
    return mapping


# ---------------------------------------------------------------------------
# 3) ตัวช่วยแปลงค่า
# ---------------------------------------------------------------------------
def num(value):
    if value is None:
        return None
    s = str(value).strip()
    if s.lower() in NA_VALUES:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def n0(value):
    v = num(value)
    return 0.0 if v is None else v


def clean_gum(raw):
    """
    ทำความสะอาดค่า gum ให้เป็น 6 sextant
      - ค่าสั้นกว่า 6 หลัก ตีความเป็นค่าเดียวกันทั้งปาก ('9' -> '999999')
      - เลข 4 และ 5 (รหัส CPI เดิม) แปลงเป็น 3 (ปริทันต์อักเสบ)
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if s.lower() in NA_VALUES:
        return None
    s = s.split(".")[0]
    if not s.isdigit():
        return None
    if len(s) == 1:
        s = s * 6
    if len(s) != 6:
        return None
    out = []
    for ch in s:
        d = int(ch)
        if d in (4, 5):
            d = 3
        if d not in (0, 1, 2, 3, 9):
            return None
        out.append(d)
    return out


def gum_status(sextants):
    valid = [d for d in sextants if d in (0, 1, 2, 3)]
    return 9 if not valid else max(valid)


def new_bucket():
    return {
        "total": 0, "examined": 0, "examinedOther": 0, "quality": 0,
        "cariesFree": 0, "cariesUntreated": 0, "extracted": 0, "filled": 0, "dmftAny": 0,
        "sumTeeth": 0.0, "sumCaries": 0.0, "sumExtract": 0.0, "sumFilling": 0.0,
        "gumN": 0, "gum0": 0, "gum1": 0, "gum2": 0, "gum3": 0, "gum9": 0, "gumBad": 0,
        "sxt0": 0, "sxt1": 0, "sxt2": 0, "sxt3": 0, "sxt9": 0,
        "func20": 0, "occl4": 0, "func20occl4": 0,
        "needFluoride": 0, "needSealant": 0, "needFilling": 0, "needExtract": 0, "needScaling": 0,
        "failProvider": 0, "failRange": 0,
    }


# ---------------------------------------------------------------------------
# 4) ปีงบประมาณ
# ---------------------------------------------------------------------------
def find_fiscal_year():
    """
    ปีงบประมาณไทย 1 ต.ค. ถึง 30 ก.ย. กำหนดจากวันที่ตรวจล่าสุดในข้อมูล
        ล่าสุด 7 ก.ค. 2569  -> ปีงบประมาณ 2569
        ล่าสุด 7 ต.ค. 2569  -> ปีงบประมาณ 2570
    """
    latest = None
    if META.exists():                                      # บันทึกไว้แล้วตอนรวมไฟล์
        try:
            latest = json.loads(META.read_text(encoding="utf-8")).get("latestServe")
        except Exception:
            latest = None
    if latest is None:                                     # กรณีใช้ combined.csv เดิมที่ไม่มี meta
        with open(COMBINED, encoding="utf-8-sig", newline="") as fh:
            for row in csv.DictReader(fh):
                d = (row.get("date_serv") or "").strip()
                if d.lower() in NA_VALUES:
                    continue
                if latest is None or d > latest:
                    latest = d
    if latest is None:
        raise UserError(
            "ไม่พบผู้ที่ได้รับการตรวจฟันเลยในข้อมูลชุดนี้",
            ["ทุกรายการมีคอลัมน์ date_serv ว่าง หมายความว่ายังไม่มีใครได้รับการตรวจ",
             "กรุณาตรวจสอบว่าส่งออกข้อมูลจาก HDC ถูกช่วงเวลาหรือไม่"])
    year, month = int(latest[:4]), int(latest[5:7])
    end_year = year + 1 if month >= 10 else year
    return end_year + 543, f"{end_year - 1}-10-01", f"{end_year}-09-30", latest


# ---------------------------------------------------------------------------
# 5) คำนวณสรุปราย hoscode
# ---------------------------------------------------------------------------
def summarise(hospitals, fy):
    fy_be, fy_start, fy_end, latest = fy
    buckets = defaultdict(new_bucket)
    units = {}
    unknown_codes = set()
    rows_total = 0

    with open(COMBINED, encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            rows_total += 1
            age = (row.get("agesurvey") or "").strip()
            if age not in AGE_LABEL:
                continue
            hoscode = (row.get("hoscode") or "").strip()
            b = buckets[(age, hoscode)]
            b["total"] += 1

            if hoscode not in units:
                info = hospitals.get(hoscode)
                if info is None:
                    unknown_codes.add(hoscode)
                    info = {}
                units[hoscode] = {
                    "hoscode": hoscode,
                    "hosname": info.get("hosname") or (row.get("hosname") or "").strip() or hoscode,
                    "ampname": info.get("ampname") or "ไม่ทราบอำเภอ",
                    "pvname": info.get("pvname") or "ไม่ทราบจังหวัด",
                }

            date_serv = (row.get("date_serv") or "").strip()
            if date_serv.lower() in NA_VALUES:
                continue
            if not (fy_start <= date_serv <= fy_end):
                b["examinedOther"] += 1
                continue
            b["examined"] += 1

            # ---------- เกณฑ์คุณภาพ ----------
            provider_ok = (row.get("providertype") or "").strip() in ("02", "06")
            pteeth, pcaries = num(row.get("pteeth")), num(row.get("pcaries"))
            pfilling, pextract = num(row.get("pfilling")), num(row.get("pextract"))
            dteeth, dcaries = num(row.get("dteeth")), num(row.get("dcaries"))
            dfilling, dextract = num(row.get("dfilling")), num(row.get("dextract"))
            pp, ppr = num(row.get("permanent_permanent")), num(row.get("permanent_prosthesis"))
            prpr = num(row.get("prosthesis_prosthesis"))

            if age == "3ปี":
                range_ok = (
                    None not in (dteeth, dcaries, dfilling, dextract)
                    and 1 <= dteeth <= 20
                    and dcaries + dfilling + dextract <= 20
                    and dteeth + dextract == 20)
            elif age in ("6ปี", "12ปี"):
                limit = 12 if age == "6ปี" else 28
                range_ok = (
                    None not in (pteeth, pcaries, pfilling, pextract)
                    and 1 <= pteeth <= limit
                    and pfilling + pextract + pcaries <= limit
                    and pcaries + pfilling <= pteeth)
            else:
                range_ok = (
                    None not in (pteeth, pextract, pp, ppr, prpr)
                    and 0 <= pteeth <= 32
                    and 0 <= pp + ppr + prpr <= 10
                    and pteeth + pextract > 0)

            if not provider_ok:
                b["failProvider"] += 1
            elif not range_ok:
                b["failRange"] += 1
            if not (provider_ok and range_ok):
                continue
            b["quality"] += 1

            # ---------- ตารางที่ 2 และ 3 ----------
            if age == "3ปี":
                teeth, caries, filling, extract = n0(dteeth), n0(dcaries), n0(dfilling), n0(dextract)
            else:
                teeth, caries, filling, extract = n0(pteeth), n0(pcaries), n0(pfilling), n0(pextract)

            if caries + filling + extract == 0:
                b["cariesFree"] += 1
            else:
                b["dmftAny"] += 1
            b["cariesUntreated"] += int(caries > 0)
            b["extracted"] += int(extract > 0)
            b["filled"] += int(filling > 0)
            b["sumTeeth"] += teeth
            b["sumCaries"] += caries
            b["sumExtract"] += extract
            b["sumFilling"] += filling

            # ---------- ตารางที่ 4 และ 5 ----------
            if age in ("12ปี", "60ปี"):
                sextants = clean_gum(row.get("gum"))
                if sextants is None:
                    b["gumBad"] += 1
                else:
                    b["gumN"] += 1
                    b[f"gum{gum_status(sextants)}"] += 1
                    for d in sextants:
                        b[f"sxt{d}"] += 1

            # ---------- ตารางที่ 6 ----------
            if age == "60ปี":
                f20 = n0(pteeth) >= 20
                o4 = n0(pp) + n0(ppr) + n0(prpr) >= 4
                b["func20"] += int(f20)
                b["occl4"] += int(o4)
                b["func20occl4"] += int(f20 and o4)

            # ---------- ตารางที่ 7 ----------
            b["needFluoride"] += int((row.get("need_fluoride") or "").strip() == "1")
            b["needScaling"] += int((row.get("need_scaling") or "").strip() == "1")
            b["needSealant"] += int(n0(row.get("need_sealant")) > 0)
            if age == "3ปี":
                b["needFilling"] += int(n0(row.get("need_dfilling")) > 0)
                b["needExtract"] += int(n0(row.get("need_dextract")) > 0)
            else:
                b["needFilling"] += int(n0(row.get("need_pfilling")) > 0)
                b["needExtract"] += int(n0(row.get("need_pextract")) > 0)

    if unknown_codes:
        say(f"   ! รหัสหน่วยบริการ {len(unknown_codes)} รหัสไม่มีในตาราง hospitals.csv "
            f"จัดไว้กลุ่ม 'ไม่ทราบอำเภอ'")

    records = []
    for (age, hoscode), b in sorted(buckets.items()):
        rec = {"age": age, "hoscode": hoscode}
        rec.update({k: (round(v, 2) if isinstance(v, float) else v) for k, v in b.items()})
        records.append(rec)

    return {
        "app": APP_NAME,
        "version": VERSION,
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "rowsTotal": rows_total,
        "fiscalYear": fy_be,
        "fyStart": fy_start,
        "fyEnd": fy_end,
        "latestServe": latest,
        "ageOrder": AGE_ORDER,
        "ageLabel": AGE_LABEL,
        "units": [units[k] for k in sorted(units)],
        "records": records,
    }


# ---------------------------------------------------------------------------
def version_key(text):
    return tuple(int(x) for x in re.findall(r"\d+", text or "")[:4]) or (0,)


def check_update():
    """
    ถามเลขเวอร์ชันล่าสุดจาก GitHub เพื่อแจ้งผู้ใช้ ไม่ได้ส่งข้อมูลใด ๆ ของพื้นที่ออกไป
    ถ้าไม่มีอินเทอร์เน็ต ถูกไฟร์วอลล์กั้น หรือช้าเกิน 3 วินาที จะข้ามไปเงียบ ๆ
    """
    if not GITHUB_REPO:
        return
    try:
        import urllib.request
        req = urllib.request.Request(
            f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest",
            headers={"User-Agent": "OralHealthDashboard", "Accept": "application/vnd.github+json"})
        with urllib.request.urlopen(req, timeout=3) as resp:
            tag = json.loads(resp.read().decode("utf-8")).get("tag_name", "")
    except Exception:
        return
    if not tag or version_key(tag) <= version_key(VERSION):
        return
    say("")
    say(f"  * มีโปรแกรมเวอร์ชันใหม่แล้ว: {tag}  (ที่ใช้อยู่คือ {VERSION})")
    say(f"    ดาวน์โหลดได้ที่ https://github.com/{GITHUB_REPO}/releases/latest")
    say(f"    วิธีอัปเดตคือนำไฟล์ .exe ใหม่มาวางทับไฟล์เดิม ข้อมูลในโฟลเดอร์ data ยังอยู่ครบ")


class UserError(Exception):
    """ข้อผิดพลาดที่อธิบายให้ผู้ใช้เข้าใจได้ ไม่ต้องแสดง traceback"""

    def __init__(self, title, hints=None):
        super().__init__(title)
        self.title = title
        self.hints = hints or []


# ---------------------------------------------------------------------------
def run(argv):
    force = "--force" in argv
    full_csv = "--full-csv" in argv
    no_open = "--no-open" in argv

    hr()
    say(f"  {APP_NAME}   เวอร์ชัน {VERSION}")
    hr()
    say(f"โฟลเดอร์ทำงาน : {APP_DIR}")
    say(f"ผลลัพธ์        : {OUT_DIR}")
    if "--no-update-check" not in argv:
        check_update()
    say("")

    if ensure_data_folders():
        say("สร้างโฟลเดอร์ data ให้เรียบร้อยแล้ว")
        raise UserError(
            "ยังไม่มีข้อมูลสำหรับประมวลผล",
            ["กรุณานำไฟล์ Excel ที่ส่งออกจากระบบ HDC มาวางในโฟลเดอร์",
             *[f"    {DATA_DIR / a}" for a in AGE_ORDER],
             "แล้วเปิดโปรแกรมนี้อีกครั้ง"])

    say("[1/4] อ่านและรวมไฟล์ Excel")
    combine_excels(force=force, full_csv=full_csv)

    say("")
    say("[2/4] อ่านตารางหน่วยบริการ")
    hospitals = load_hospitals()

    say("")
    say("[3/4] คำนวณผลตามเงื่อนไขรายงาน")
    fy = find_fiscal_year()
    say(f"   + วันที่ตรวจล่าสุดในข้อมูล {fy[3]}  →  ปีงบประมาณ {fy[0]} "
        f"(ค.ศ. {fy[1]} ถึง {fy[2]})")
    summary = summarise(hospitals, fy)
    SUMMARY.write_text(json.dumps(summary, ensure_ascii=False), encoding="utf-8")

    ex = sum(r["examined"] for r in summary["records"])
    qa = sum(r["quality"] for r in summary["records"])
    tt = sum(r["total"] for r in summary["records"])
    other = sum(r["examinedOther"] for r in summary["records"])
    if ex == 0:
        raise UserError(
            f"ไม่พบผู้ที่ได้รับการตรวจฟันในปีงบประมาณ {fy[0]}",
            [f"มีผู้ได้รับการตรวจนอกปีงบประมาณนี้ {other:,} รายการ",
             "กรุณาตรวจสอบช่วงเวลาที่ส่งออกข้อมูลจาก HDC"])
    say(f"   + ประชากรทั้งหมด {tt:,} คน")
    say(f"   + ได้รับการตรวจฟันในปีงบ {fy[0]}: {ex:,} คน ({ex / tt * 100:.1f}%)")
    say(f"   + ผ่านเกณฑ์คุณภาพ {qa:,} คน ({qa / ex * 100:.1f}% ของผู้ที่ตรวจ)")
    say(f"   + ตรวจนอกปีงบประมาณ ไม่นำมาคำนวณ {other:,} รายการ")
    say(f"   + หน่วยบริการทั้งหมด {len(summary['units'])} แห่ง")

    say("")
    say("[4/4] สร้างรายงาน dashboard.html")
    template = resource("dashboard_template.html")
    if template is None:
        raise UserError("ไม่พบไฟล์แม่แบบ dashboard_template.html",
                        ["ถ้าใช้งานแบบไฟล์ .exe แสดงว่าไฟล์ .exe เสียหาย กรุณาดาวน์โหลดใหม่"])
    html = template.read_text(encoding="utf-8")
    html = html.replace("/*__DATA__*/null", json.dumps(summary, ensure_ascii=False))
    OUTPUT.write_text(html, encoding="utf-8")
    say(f"   = {OUTPUT}")

    hr()
    say("  เสร็จเรียบร้อย")
    hr()

    if not no_open:
        try:
            webbrowser.open(OUTPUT.as_uri())
            say("กำลังเปิดรายงานในเบราว์เซอร์ ...")
        except Exception:
            say(f"เปิดเบราว์เซอร์อัตโนมัติไม่ได้ กรุณาเปิดไฟล์นี้เอง: {OUTPUT}")


def main():
    global _log_fh
    setup_console()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    try:
        _log_fh = open(LOGFILE, "w", encoding="utf-8")
        _log_fh.write(f"{APP_NAME} {VERSION} — {datetime.now():%Y-%m-%d %H:%M:%S}\n")
    except Exception:
        _log_fh = None

    code = 0
    try:
        run(sys.argv[1:])
    except UserError as e:
        say("")
        hr()
        say(f"  ไม่สามารถทำงานต่อได้: {e.title}")
        hr()
        for h in e.hints:
            say(f"  {h}")
        code = 1
    except PermissionError as e:
        say("")
        hr()
        say("  ไม่สามารถเขียนไฟล์ได้")
        hr()
        say(f"  {e}")
        say("  สาเหตุที่พบบ่อยคือเปิดไฟล์ผลลัพธ์ค้างไว้ใน Excel หรือเบราว์เซอร์")
        say("  กรุณาปิดไฟล์นั้นแล้วเปิดโปรแกรมอีกครั้ง")
        code = 1
    except Exception:
        say("")
        hr()
        say("  เกิดข้อผิดพลาดที่ไม่คาดคิด")
        hr()
        say(traceback.format_exc())
        say(f"  กรุณาส่งไฟล์ {LOGFILE} ให้ผู้ดูแลระบบเพื่อตรวจสอบ")
        code = 2
    finally:
        if _log_fh:
            try:
                _log_fh.close()
            except Exception:
                pass
    pause()
    return code


if __name__ == "__main__":
    sys.exit(main())
